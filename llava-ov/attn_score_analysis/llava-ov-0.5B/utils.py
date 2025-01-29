import os
import json
import argparse
import pandas as pd
from sklearn.metrics import confusion_matrix
from openpyxl import load_workbook
from openpyxl.comments import Comment


# 融合不同的generate文件
def fun1():
    all_file_path = [
        f"/cpfs/user/wangpengfei2/project/LLaVA-NeXT/attn_score_analysis/{args.model_name}/{args.task_type}/info/pred0.json",
    ]
    all_dir_path = [
        f"/cpfs/user/wangpengfei2/project/LLaVA-NeXT/attn_score_analysis/{args.model_name}/{args.task_type}/bench_res{num}"
            for num in range(10)
    ]
    all_res = []
    for file_path in all_file_path:
        all_res += json.loads(open(file_path,'r',encoding="utf-8").read())

    for dir_path in all_dir_path:
        res_files = [os.path.join(dir_path, x) for x in os.listdir(dir_path)]
        for res in res_files:
            all_res += json.loads(open(res,'r',encoding="utf-8").read())

    # 去重后排序
    all_res = list({(item['dataset_name'], item['sample_id']): item for item in all_res if item["answer"] is not None}.values())
    all_res.sort(key=lambda x: (x['dataset_name'], x['sample_id']))
    for item in all_res:
        score_path = os.path.join(
            args.root_path, "scores",
            f'{item["dataset_name"]}_sample{item["sample_id"]}_imgnum{len(item["raw_img_list"])}_target{item["target_id"]}_scores.xlsx'
        )
        if not os.path.exists(score_path):
            print({item["dataset_name"]}, {item["sample_id"]})
    print(f"total pred result number: {len(all_res)}")

    open(os.path.join(args.root_path, "info/pred.json"),'w',encoding="utf-8").write(json.dumps(all_res, indent=4))

# 融合不同的analysis结果文件
def fun2():
    marged_info = f"/cpfs/user/wangpengfei2/project/usage/attn_score_analysis/make_data/merged_result_{args.task_type}.json"
    marged_info = json.loads(open(marged_info,'r',encoding="utf-8").read())
    marged_info = {(item['dataset_name'], item['sample_id']): item for item in marged_info}

    pred = os.path.join(args.root_path, "info/pred.json")
    pred = json.loads(open(pred,'r',encoding="utf-8").read())
    pred = {(item['dataset_name'], item['sample_id']): item for item in pred}

    is_right = os.path.join(args.root_path, "info/eval_score.json")
    is_right = json.loads(open(is_right,'r',encoding="utf-8").read())
    is_right = {(item['dataset_name'], item['sample_id']): item for item in is_right}

    scores_path = os.path.join(args.root_path, "scores")
    scores = os.listdir(scores_path)
    real_scores = [os.path.join(scores_path, x) for x in scores]
    
    all_res = []
    used = []
    for i in range(len(real_scores)):
        file_name = scores[i].split("_")
        dataset_name = file_name[0]
        sample_id = int(file_name[1].strip("sample"))
        if (dataset_name, sample_id) not in pred or (dataset_name, sample_id) not in marged_info:
            # print((dataset_name, sample_id))
            continue
        if (dataset_name, sample_id) in used:
            print()
        used.append((dataset_name, sample_id))
        item = pred[(dataset_name, sample_id)]
        df = pd.read_excel(real_scores[i], sheet_name="total_info")
        row_index = df[df.iloc[:, 0] == f"decoder {total_layer}"].index[0]
        row_data = df.iloc[row_index]
        max_col = row_data.iloc[1:].idxmax()
        item["gpt_answer"] = marged_info[(dataset_name, sample_id)]["answer"] if "answer" in marged_info[(dataset_name, sample_id)] else ""
        item["is_right"] = is_right[(dataset_name, sample_id)]["score"]
        item["attn_img_num"] = int(max_col.strip("image"))
        all_res.append(item)
    all_res.sort(key=lambda x: (x['dataset_name'], x['sample_id']))
    print(f"total data number: {len(marged_info)}")
    print(f"total pred result number: {len(pred)}")
    print(f"final result number: {len(all_res)}")
    open(os.path.join(args.root_path, "info/analysis.json"),'w',encoding="utf-8").write(json.dumps(all_res, indent=4))


# 计算混淆矩阵
def cal_matrix():
    analysis = os.path.join(args.root_path, "info/analysis.json")
    analysis = json.loads(open(analysis,'r',encoding="utf-8").read())

    is_right = []
    attn_index = []
    for item in analysis:
        # if item["dataset_name"] != "TQA":
        is_right.append(item["is_right"])
        attn_index.append(int(item["attn_img_num"]==item["target_id"]))
        # if is_right[-1] == 0 and attn_index[-1] == 0:
        #     print(item["dataset_name"], item["sample_id"])
            # if len(item["raw_img_list"]) == 2:
            #     cnt += 1
    # print(cnt)

    cm = confusion_matrix(is_right, attn_index)
    print(cm)
    precision_0 = cm[0, 1] / (cm[0, 0] + cm[0, 1])  # 类别 0 的精度
    precision_1 = cm[1, 1] / (cm[1, 1] + cm[1, 0])  # 类别 1 的精度
    recall_0 = cm[1, 0] / (cm[0, 0] + cm[1, 0])     # 类别 0 的召回率
    recall_1 = cm[1, 1] / (cm[1, 1] + cm[0, 1])     # 类别 1 的召回率
    total_num = cm[0, 0] + cm[0, 1] + cm[1, 1] + cm[1, 0]

    # 将混淆矩阵转化为 DataFrame 以便写入 Excel
    df_cm = pd.DataFrame(cm, index=["answer_false", "answer_true"], columns=["attn_false", "attn_true"])
    df_cm["Precision"] = [precision_0, precision_1]
    df_cm.loc["Recall"] = [recall_0, recall_1, ""]
    df_cm[''] = ["", "", ""]
    df_cm['Percentage1'] = [cm[0, 0] / total_num, cm[1, 0] / total_num, ""]
    df_cm['Percentage2'] = [cm[0, 1] / total_num, cm[1, 1] / total_num, ""]

    # 将混淆矩阵写入 Excel
    output_file = os.path.join(args.root_path, f"info/confusion_matrix_{args.model_name}_{args.task_type}data.xlsx")
    df_cm.to_excel(output_file, sheet_name="Confusion Matrix")

    # 加载生成的 Excel 文件并准备添加注释
    wb = load_workbook(output_file)
    ws = wb['Confusion Matrix']

    # 添加注释
    comments = {
        (2, 2): "True Negative: 回答错误，attn错误",
        (2, 3): "False Positive: 回答错误，attn正确",
        (3, 2): "False Negative: 回答正确，attn错误",
        (3, 3): "True Positive: 回答正确，attn正确"
    }

    for cell, text in comments.items():
        ws.cell(row=cell[0], column=cell[1]).comment = Comment(text, args.model_name)

    # 保存带有注释的 Excel 文件
    wb.save(output_file)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--function", type=list, default=[2,3])
    parser.add_argument("--root_path", type=str, default="/cpfs/user/wangpengfei2/project/LLaVA-NeXT/attn_score_analysis")
    parser.add_argument("--model_name", type=str, default="llava-ov-0.5B")
    parser.add_argument("--task_type", type=str, default="medium", choices=["easy", "medium", "hard"])

    args = parser.parse_args()
    args.root_path = os.path.join(args.root_path, args.model_name, args.task_type)
    os.makedirs(os.path.join(args.root_path, "info"), exist_ok=True)
    total_layer = 24

    for fun_num in args.function:
        if int(fun_num) == 1:
            fun1()
        elif int(fun_num) == 2:
            fun2()
        elif int(fun_num) == 3:
            cal_matrix()
    print(len(os.listdir(f"/cpfs/user/wangpengfei2/project/LLaVA-NeXT/attn_score_analysis/{args.model_name}/{args.task_type}/scores")))
